"""
Product Import Service
Handles multi-source product imports with source-aware mapping
"""
from typing import Dict, Optional, List, Any
from pathlib import Path
import openpyxl
from slugify import slugify


def normalize_ean(raw: Any) -> Optional[str]:
    """Return digits-only EAN string or None if empty/invalid."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits or None


def is_valid_ean13(ean: Optional[str]) -> bool:
    """Check EAN has exactly 13 digits."""
    return bool(ean) and ean.isdigit() and len(ean) == 13


class SourceMapper:
    """Base class for source-specific column mapping"""
    
    def __init__(self):
        self.source_name = "unknown"
        self.header_row = 1  # Default: first row is header
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        """
        Map source-specific columns to standard product fields
        Returns None if row should be skipped
        """
        raise NotImplementedError
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        """Extract and validate EAN from row"""
        raise NotImplementedError


class EffezzetaMapper(SourceMapper):
    """Mapper for Effezzeta (Listino-prodotti.xlsx)"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "effezzeta"
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        return normalize_ean(row.get("Codice a barre EAN-13 o JAN"))
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title
        title = row.get("Nome prodotto")
        if not title or not str(title).strip():
            return None
        
        # Extract price (IVA Esclusa)
        price = row.get("Listino - IVA Esclusa")
        try:
            price_value = float(price) if price else None
        except (ValueError, TypeError):
            price_value = None
        
        # Extract stock
        stock = row.get("Quantità")
        try:
            stock_value = int(stock) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract category (single level)
        category = row.get("Categoria")
        category_path = [str(category).strip()] if category and str(category).strip() else []
        
        # Extract description
        description = row.get("Descrizione")
        description_text = str(description).strip() if description else None
        
        # Extract images (comma-separated URLs)
        images_str = row.get("URL di immagini prodotto")
        image_urls = []
        if images_str and str(images_str).strip():
            urls = str(images_str).split(',')
            image_urls = [url.strip() for url in urls if url.strip()]
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": price_value,
            "stock": stock_value,
            "brand_name": None,  # Effezzeta doesn't have brand
            "category_path": category_path,
            "description": description_text,
            "image_urls": image_urls,
            "source": self.source_name,
            "row_number": row_number
        }


class ErregameMapper(SourceMapper):
    """Mapper for Erregame (erregame_organized.xlsx)"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "erregame"
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        return normalize_ean(row.get("EAN"))
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title
        title = row.get("Title")
        if not title or not str(title).strip():
            return None
        
        # Extract price (final price)
        price = row.get("Price")
        try:
            price_value = float(price) if price else None
        except (ValueError, TypeError):
            price_value = None
        
        # Extract stock
        stock = row.get("Available")
        try:
            stock_value = int(stock) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract brand
        brand = row.get("Brand")
        brand_name = str(brand).strip() if brand and str(brand).strip() else None
        
        # Extract category path (Category → Subcategory)
        category = row.get("Category")
        subcategory = row.get("Subcategory")
        category_path = []
        if category and str(category).strip():
            category_path.append(str(category).strip())
            if subcategory and str(subcategory).strip():
                category_path.append(str(subcategory).strip())
        
        # Extract description
        description = row.get("Description")
        description_text = str(description).strip() if description and str(description).strip() else None
        
        # Extract image URL (single)
        image_url = row.get("ImageLink")
        image_urls = [str(image_url).strip()] if image_url and str(image_url).strip() else []
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": price_value,
            "stock": stock_value,
            "brand_name": brand_name,
            "category_path": category_path,
            "description": description_text,
            "image_urls": image_urls,
            "source": self.source_name,
            "row_number": row_number
        }


class DixeMapper(SourceMapper):
    """Mapper for Dixe (Dixe_organized.xlsx)"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "dixe"
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        return normalize_ean(row.get("COD/EAN"))
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title
        title = row.get("Titolo")
        if not title or not str(title).strip():
            return None
        
        # Extract stock
        stock = row.get("quantity")
        try:
            stock_value = int(stock) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract category path (format: "Parent > Child")
        category = row.get("Categoria")
        category_path = []
        if category and str(category).strip():
            parts = str(category).split('>')
            category_path = [p.strip() for p in parts if p.strip()]
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": None,  # Dixe doesn't have price
            "stock": stock_value,
            "brand_name": None,  # Dixe doesn't have brand
            "category_path": category_path,
            "description": None,
            "image_urls": [],
            "source": self.source_name,
            "row_number": row_number
        }


class TelefoniaMapper(SourceMapper):
    """Mapper for Listino Telefonia web.xlsx"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "telefonia"
        self.header_row = 7  # Header starts at row 7
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        # Read EAN directly from "EAN" column (column H)
        ean = row.get("EAN")
        if ean:
            # Convert to string to preserve leading zeros
            ean_str = str(ean).strip()
            # Remove .0 suffix if pandas read it as float
            if ean_str.endswith('.0'):
                ean_str = ean_str[:-2]
            return ean_str if len(ean_str) == 13 and ean_str.isdigit() else None
        return None
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title from "TELEFONI CELLULARI" column
        title = row.get("TELEFONI CELLULARI")
        if not title or not str(title).strip():
            return None
        
        # Extract price from "PREZZO REVERSE CHARGE"
        price = row.get("PREZZO REVERSE CHARGE")
        try:
            price_value = float(price) if price else None
        except (ValueError, TypeError):
            price_value = None
        
        # Extract stock from "DISP"
        stock = row.get("DISP")
        try:
            stock_value = int(stock) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract brand from "MARCA"
        brand = row.get("MARCA")
        brand_name = str(brand).strip() if brand and str(brand).strip() else None
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": price_value,
            "stock": stock_value,
            "brand_name": brand_name,
            "category_path": ["Telefonia"],
            "description": None,
            "image_urls": [],
            "source": self.source_name,
            "row_number": row_number
        }


class InformaticaMapper(SourceMapper):
    """Mapper for Listino INFORMATICA web.xlsx"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "informatica"
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        # Need to check actual structure - seems complex
        return None
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        # TODO: Complex structure - needs manual inspection
        return None


class GiochiMapper(SourceMapper):
    """Mapper for Listino GIOCHI.xlsx"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "giochi"
        self.header_row = 9  # Header starts at row 9
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        # Read EAN directly from "EAN" column (column G, row 9 header)
        ean = row.get("EAN")
        if ean:
            # Convert to string to preserve leading zeros
            ean_str = str(ean).strip()
            # Remove .0 suffix if pandas read it as float
            if ean_str.endswith('.0'):
                ean_str = ean_str[:-2]
            return ean_str if len(ean_str) == 13 and ean_str.isdigit() else None
        return None
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title from "CONSOLE" column
        title = row.get("CONSOLE")
        if not title or not str(title).strip():
            return None
        
        # Extract stock from "DISP."
        stock = row.get("DISP.")
        try:
            stock_value = int(stock) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract brand from "MARCA"
        brand = row.get("MARCA")
        brand_name = str(brand).strip() if brand and str(brand).strip() else None
        
        # Extract category from "CATEGORIA"
        category = row.get("CATEGORIA")
        category_path = []
        if category and str(category).strip():
            parts = str(category).split('/')
            category_path = [p.strip() for p in parts if p.strip()]
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": None,  # No price column
            "stock": stock_value,
            "brand_name": brand_name,
            "category_path": category_path if category_path else ["Giochi"],
            "description": None,
            "image_urls": [],
            "source": self.source_name,
            "row_number": row_number
        }


class CartoleriaMapper(SourceMapper):
    """Mapper for Listino Cartoleria.xlsx"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "cartoleria"
        self.header_row = 4  # Header starts at row 4
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        # Read EAN directly from "EAN" column (column 8, row 4 header)
        ean = row.get("EAN")
        if ean:
            # Convert to string to preserve leading zeros
            ean_str = str(ean).strip()
            # Remove .0 suffix if pandas read it as float
            if ean_str.endswith('.0'):
                ean_str = ean_str[:-2]
            return ean_str if len(ean_str) == 13 and ean_str.isdigit() else None
        return None
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title from "DESCRIZIONE" column
        title = row.get("DESCRIZIONE")
        if not title or not str(title).strip():
            return None
        
        # Extract stock from "Disp."
        stock = row.get("Disp.")
        try:
            stock_value = int(float(stock)) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract price from "prezzo iva esclusa"
        price = row.get("prezzo iva esclusa")
        try:
            price_value = float(price) if price else None
        except (ValueError, TypeError):
            price_value = None
        
        # Extract brand from "MARCA"
        brand = row.get("MARCA")
        brand_name = str(brand).strip() if brand and str(brand).strip() else None
        
        # Extract category from "CATEGORIA"
        category = row.get("CATEGORIA")
        category_path = []
        if category and str(category).strip():
            parts = str(category).split('/')
            category_path = [p.strip() for p in parts if p.strip()]
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": price_value,
            "stock": stock_value,
            "brand_name": brand_name,
            "category_path": category_path if category_path else ["Cartoleria"],
            "description": None,
            "image_urls": [],
            "source": self.source_name,
            "row_number": row_number
        }


class AccessoriMapper(SourceMapper):
    """Mapper for Listino ACCESSORI telefonia.xlsx"""
    
    def __init__(self):
        super().__init__()
        self.source_name = "accessori"
        self.header_row = 6  # Header starts at row 6
    
    def get_ean(self, row: Dict[str, Any]) -> Optional[str]:
        # Read EAN directly from "EAN" column (column 7, row 6 header)
        ean = row.get("EAN")
        if ean:
            # Convert to string to preserve leading zeros
            ean_str = str(ean).strip()
            # Remove .0 suffix if pandas read it as float
            if ean_str.endswith('.0'):
                ean_str = ean_str[:-2]
            return ean_str if len(ean_str) == 13 and ean_str.isdigit() else None
        return None
    
    def map_row(self, row: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
        ean = self.get_ean(row)
        if not ean:
            return None
        
        # Extract title from "MODELLO " column (note the space)
        title = row.get("MODELLO ") or row.get("MODELLO")
        if not title or not str(title).strip():
            return None
        
        # Extract stock from "DISPON."
        stock = row.get("DISPON.")
        try:
            stock_value = int(float(stock)) if stock else 0
        except (ValueError, TypeError):
            stock_value = 0
        
        # Extract price from "prezzo         Iva  esclusa" (with spaces)
        price = row.get("prezzo         Iva  esclusa") or row.get("prezzo Iva esclusa")
        try:
            price_value = float(price) if price else None
        except (ValueError, TypeError):
            price_value = None
        
        # Extract brand from "MARCA"
        brand = row.get("MARCA")
        brand_name = str(brand).strip() if brand and str(brand).strip() else None
        
        # Extract category from "ACCESSORI TELEFONIA"
        category = row.get("ACCESSORI TELEFONIA")
        category_path = []
        if category and str(category).strip():
            parts = str(category).split('/')
            category_path = [p.strip() for p in parts if p.strip()]
        
        return {
            "ean": ean,
            "title": str(title).strip(),
            "price": price_value,
            "stock": stock_value,
            "brand_name": brand_name,
            "category_path": category_path if category_path else ["Accessori Telefonia"],
            "description": None,
            "image_urls": [],
            "source": self.source_name,
            "row_number": row_number
        }


class ProductImportService:
    """Service for importing products from Excel files"""
    
    MAPPERS = {
        "effezzeta": EffezzetaMapper,
        "erregame": ErregameMapper,
        "dixe": DixeMapper,
        "telefonia": TelefoniaMapper,
        "informatica": InformaticaMapper,
        "giochi": GiochiMapper,
        "cartoleria": CartoleriaMapper,
        "accessori": AccessoriMapper,
    }
    
    CHUNK_SIZE = 300
    
    def __init__(self, source: str):
        if source not in self.MAPPERS:
            raise ValueError(f"Unknown source: {source}")
        
        self.source = source
        self.mapper = self.MAPPERS[source]()
    
    def read_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Read Excel file and return list of row dictionaries"""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Get header row from mapper configuration
        header_row = self.mapper.header_row
        
        # Get headers from specified row
        headers = []
        for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
            headers.append(cell.value)
        
        # Read data rows (starting from row after header)
        rows = []
        data_start_row = header_row + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            row_dict = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    row_dict[headers[idx]] = value
            rows.append({"data": row_dict, "row_number": row_idx})
        
        return rows
    
    def map_rows(self, rows: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Map raw rows to standard format
        Returns: (valid_rows, skipped_rows)
        """
        valid_rows = []
        skipped_rows = []
        
        for row_info in rows:
            row_data = row_info["data"]
            row_number = row_info["row_number"]

            ean = self.mapper.get_ean(row_data)
            if not ean:
                skipped_rows.append({
                    "row_number": row_number,
                    "ean": None,
                    "reason": "missing_ean",
                    "details": "Product has no EAN code"
                })
                continue

            if not is_valid_ean13(ean):
                skipped_rows.append({
                    "row_number": row_number,
                    "ean": ean,
                    "reason": "invalid_ean13",
                    "details": "EAN must be exactly 13 digits"
                })
                continue

            mapped = self.mapper.map_row(row_data, row_number)

            if mapped is None:
                skipped_rows.append({
                    "row_number": row_number,
                    "ean": ean,
                    "reason": "missing_title",
                    "details": "Product has no title"
                })
            else:
                valid_rows.append(mapped)
        
        return valid_rows, skipped_rows
    
    def chunk_rows(self, rows: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
        """Split rows into chunks for batch processing"""
        chunks = []
        for i in range(0, len(rows), self.CHUNK_SIZE):
            chunks.append(rows[i:i + self.CHUNK_SIZE])
        return chunks
