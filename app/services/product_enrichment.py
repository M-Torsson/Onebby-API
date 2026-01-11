"""Product enrichment import (fill missing fields using EAN as key)."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl


class EnrichmentReader:
    """Read vendor XLS/XLSX files and normalize rows for enrichment.

    The files in app/excel contain varying header positions; we detect the header row
    by the first row that contains a cell with the substring "ean" (case-insensitive).
    Only the columns we care about are extracted (ean, title, description, price, brand, images).
    """

    TITLE_HINTS = ("modello", "descrizione", "description", "telefoni", "console", "titolo")
    PRICE_HINTS = ("prezzo", "price")
    BRAND_HINTS = ("marca", "brand")
    EAN_HINTS = ("ean",)
    IMAGE_HINTS = ("image", "url", "photo", "foto")

    def __init__(self, file_path: Path):
        self.file_path = file_path

    def _normalize_str(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _normalize_ean(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        try:
            # Handle float EAN (e.g., 4894461929057.0)
            if isinstance(value, float) and value.is_integer():
                value = int(value)
        except Exception:
            pass
        text = str(value).strip()
        return text or None

    def _detect_header(self, ws) -> Tuple[int, List[str]]:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [self._normalize_str(c) for c in row]
            joined = " ".join([c.lower() for c in cells if c])
            if any(hint in joined for hint in self.EAN_HINTS):
                headers = [self._normalize_str(c) or f"col_{i+1}" for i, c in enumerate(row)]
                return row_idx, headers
            # Safety: stop scanning after 40 rows
            if row_idx >= 40:
                break
        raise ValueError("Header row with EAN column not found")

    def _find_header_index(self, headers: List[str], hints: Tuple[str, ...]) -> Optional[int]:
        for idx, name in enumerate(headers):
            if name is None:
                continue
            lower = name.lower()
            if any(h in lower for h in hints):
                return idx
        return None

    def _parse_row(self, headers: List[str], row_values: List[Any], row_number: int) -> Dict[str, Any]:
        row = {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values)))}

        ean_idx = self._find_header_index(headers, self.EAN_HINTS)
        title_idx = self._find_header_index(headers, self.TITLE_HINTS)
        desc_idx = self._find_header_index(headers, ("descr", "description"))
        price_idx = self._find_header_index(headers, self.PRICE_HINTS)
        brand_idx = self._find_header_index(headers, self.BRAND_HINTS)
        image_idx = self._find_header_index(headers, self.IMAGE_HINTS)

        def pick(idx: Optional[int]):
            return None if idx is None or idx >= len(row_values) else row_values[idx]

        ean = self._normalize_ean(pick(ean_idx))
        title = self._normalize_str(pick(title_idx))
        description = self._normalize_str(pick(desc_idx))

        price_raw = pick(price_idx)
        price: Optional[float] = None
        if price_raw not in (None, ""):
            try:
                price = float(price_raw)
            except (TypeError, ValueError):
                price = None

        brand = self._normalize_str(pick(brand_idx))

        image_urls: List[str] = []
        image_raw = pick(image_idx)
        if image_raw:
            # Split by comma to support multiple URLs in one cell
            parts = str(image_raw).split(',')
            image_urls = [p.strip() for p in parts if p and str(p).strip().lower().startswith("http")]

        # Fallback: if title missing, try first non-empty string cell
        if not title:
            for val in row_values:
                text = self._normalize_str(val)
                if text and text != ean and not text.replace('.', '', 1).isdigit():
                    title = text
                    break

        return {
            "ean": ean,
            "title": title,
            "description": description,
            "price": price,
            "brand_name": brand,
            "image_urls": image_urls,
            "row_number": row_number,
        }

    def read(self) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active

        header_row_idx, headers = self._detect_header(ws)

        rows: List[Dict[str, Any]] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not any(cell is not None for cell in row):
                continue
            mapped = self._parse_row(headers, list(row), row_idx)
            rows.append(mapped)
        return rows


__all__ = ["EnrichmentReader"]
