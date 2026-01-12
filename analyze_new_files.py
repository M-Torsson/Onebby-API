import openpyxl

print('=== تحليل الملفات الخمسة الجديدة ===\n')

files = [
    'Listino Telefonia web.xlsx',
    'Listino INFORMATICA web.xlsx',
    'Listino GIOCHI.xlsx',
    'Listino Cartoleria.xlsx',
    'Listino ACCESSORI telefonia.xlsx'
]

for filename in files:
    print(f'\n📁 {filename}')
    print('-' * 60)
    try:
        wb = openpyxl.load_workbook(f'app/excel/{filename}')
        ws = wb.active
        
        # Count non-empty rows
        data_rows = 0
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
            if any(cell.value for cell in row):
                data_rows += 1
        
        print(f'  إجمالي الصفوف: {ws.max_row}')
        print(f'  صفوف البيانات (تقريبي): {data_rows}')
        
        # Show sample rows
        print(f'\n  📋 عينة من البيانات:')
        for i in range(7, min(11, ws.max_row + 1)):
            row_data = [cell.value for cell in ws[i]][:5]
            if any(row_data):
                print(f'    Row {i}: {row_data}')
        
        wb.close()
        print(f'  ✅ الملف موجود وقابل للقراءة')
        
    except FileNotFoundError:
        print(f'  ❌ الملف غير موجود!')
    except Exception as e:
        print(f'  ⚠️ خطأ: {e}')

print('\n' + '=' * 60)
print('✅ تم فحص جميع الملفات')
print('=' * 60)
