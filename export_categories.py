# Author: Muthana
# © 2026 Muthana. All rights reserved.
# Unauthorized copying or distribution is prohibited.

"""Export Categories to Excel"""
import psycopg2  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
from datetime import datetime
import os
from dotenv import load_dotenv  # type: ignore

# Load environment variables
load_dotenv()

# Database connection
DB_URL = os.getenv('DATABASE_URL', '').replace('postgresql+psycopg2://', 'postgresql://')

def export_categories_to_excel():
    """Export all categories to a formatted Excel file"""
    
    # Connect to database
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    
    # Fetch all categories with translations
    cur.execute("""
        SELECT 
            c.id,
            c.name,
            c.slug,
            c.parent_id,
            c.is_active,
            c.sort_order,
            c.image,
            c.icon,
            c.created_at,
            c.updated_at,
            ct.lang,
            ct.name as translation_name,
            ct.description as translation_description
        FROM categories c
        LEFT JOIN category_translations ct ON c.id = ct.category_id
        ORDER BY c.id, ct.lang
    """)
    
    rows = cur.fetchall()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set headers
    headers = [
        "ID", "Name", "Slug", "Parent ID", "Active", "Sort Order",
        "Image", "Icon", "Created At", "Updated At",
        "Translation Lang", "Translation Name", "Translation Description"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # ID
        'B': 35,  # Name
        'C': 40,  # Slug
        'D': 10,  # Parent ID
        'E': 10,  # Active
        'F': 12,  # Sort Order
        'G': 60,  # Image
        'H': 30,  # Icon
        'I': 20,  # Created At
        'J': 20,  # Updated At
        'K': 12,  # Translation Lang
        'L': 35,  # Translation Name
        'M': 50,  # Translation Description
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            
            # Format dates
            if col_num in [9, 10] and value:  # Created/Updated dates
                cell.value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'strftime') else str(value)
            else:
                cell.value = str(value) if value is not None else ''
            
            # Center alignment for specific columns
            if col_num in [1, 4, 5, 6, 11]:  # ID, Active, Sort Order, Lang
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            cell.border = thin_border
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Freeze header row
    ws.freeze_panes = ws['A2']
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'categories_export_{timestamp}.xlsx'
    
    # Save file
    wb.save(filename)
    
    # Close database connection
    cur.close()
    conn.close()
    
    
    return filename


if __name__ == "__main__":
    try:
        filename = export_categories_to_excel()
        print(f"✅ Categories exported successfully to: {filename}")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
